from openpyxl import load_workbook
import sys
import gui
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog, QMainWindow

SOURCE_COLUMN = 'D'  # Location of cell value from source file.
MAX_ROWS = 500
MAX_COLUMNS = 500
TARGET_COL = 4  # In which column the cell location will be inserted at
DESCRIPTION_CELLS = ['Description', 'Manufacturer', 'MPN', 'QTY', 'CELL']


class MainWindow(gui.Ui_Dialog, QMainWindow, QDialog):
    def __init__(self, stock_path):
        super().__init__()
        self.setupUi(self)
        self.target = ""
        self.stockPath = stock_path
        self.pathOK = False
        self.fileTypeOK = False
        self.browse.clicked.connect(self.browse_files)
        self.run.clicked.connect(self.run_prog)

    def browse_files(self):
        file_name = QFileDialog.getOpenFileName(self, 'Target File', 'C:\\')
        self.path.setText(file_name[0])
        self.pathOK = True
        file_type = file_name[0].split('.')
        if file_type[1] == "xlsx":
            self.target = file_name[0]
            self.fileTypeOK = True

    def run_prog(self):
        if self.pathOK:
            if self.fileTypeOK:
                result = run(self.stockPath, self.target)
                if result:
                    self.output.setText("File Modified Successfully!")
                else:
                    self.output.setText("Could not modify target file.")
            else:
                self.output.setText("Please choose only xlsx files.")
        else:
            self.output.setText("Please choose a target file.")


def run(stock_file, bom_file):
    stock_wb = load_workbook(filename=stock_file, read_only=True)
    stock_sheetnames = stock_wb.sheetnames
    bom_wb = load_workbook(filename=bom_file)
    stock_dict = get_dict(stock_wb, stock_sheetnames)
    insert_cell_location(bom_wb, stock_dict, bom_file)
    return True


def get_dict(workbook, sheetnames):
    stk_dict = {}
    for i, sheet in enumerate(sheetnames):  # iterate through all the sheets in a wb
        workbook.active = i
        ws = workbook.active
        for row in ws.iter_rows(min_row=0, max_row=100, min_col=2, max_col=2):
            for cell in row:
                if type(cell).__name__ != 'MergedCell' and cell.value is not None and (
                        cell.value not in DESCRIPTION_CELLS):
                    cell_info_location = SOURCE_COLUMN + str(cell.row)
                    cell_location = ws[cell_info_location].value
                    stk_dict[cell.value] = cell_location
    return stk_dict


def insert_cell_location(bom_wb, stock_dict, bom_file):
    bom_ws = bom_wb.active
    bom_ws.cell(row=1, column=4).value = "CELL"  # Write a new column header "CELL"
    for row in bom_ws.iter_rows(min_row=2, max_row=100, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None and (cell.value in stock_dict.keys()):
                insert_location = SOURCE_COLUMN + str(cell.row)
                bom_ws[insert_location] = stock_dict[cell.value]
    bom_wb.save(bom_file)


if __name__ == '__main__':
    stock = "E:\\Programming\\Python\\AltiumStockSearcher\\stock.xlsx"  # TODO: change path to R&DStock.xlsx
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow(stock_path=stock)
    window.show()
    app.exec_()
